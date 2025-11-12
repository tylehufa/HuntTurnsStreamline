import pandas as pd
import os
from openpyxl import load_workbook
import numpy as np
import time
import re
from playwright.sync_api import Playwright, sync_playwright, expect, TimeoutError, Error
import datetime
from datetime import datetime, timedelta, date
import pyperclip
import pandas as pd
from pathlib import Path
import win32com.client      
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment, numbers
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
import datetime
import shutil
import pandas as pd
import logging
import traceback
import time
import tkinter as tk
from tkinter import ttk
import signal
import queue
import sys
import threading
import pythoncom
import subprocess
from concurrent.futures import ThreadPoolExecutor
import json
import pyperclip
import psutil
def signal_handler(sig, frame):
    print('\nSafely stopping the script...')
    if 'driver' in globals():
        Playwright.stop()
    sys.exit(0)

signal.signal(signal.SIGINT, signal_handler)
class ProgressBar:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("Script Progress")
        
        # Queue for thread communication
        self.queue = queue.Queue()
        
        # Window setup
        width = 400
        height = 150
        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()
        x = (screen_width/2) - (width/2)
        y = (screen_height/2) - (height/2)
        self.root.geometry('%dx%d+%d+%d' % (width, height, x, y))
        
        # Status label
        self.status_label = tk.Label(self.root, text="Initializing...", wraplength=380)
        self.status_label.pack(pady=10)
        
        # Progress bar
        self.progress = ttk.Progressbar(
            self.root,
            orient="horizontal",
            length=300,
            mode="determinate"
        )
        self.progress.pack(pady=10)
        
        # Sub-task label
        self.subtask_label = tk.Label(self.root, text="", wraplength=380)
        self.subtask_label.pack(pady=5)
        
        # Flag for thread control
        self.running = True
        
        # Start checking queue
        self.root.after(100, self.check_queue)

    def check_queue(self):
        """Check for messages from the worker thread"""
        try:
            while True:
                msg = self.queue.get_nowait()
                if msg is None:
                    self.root.quit()
                    return
                
                progress, status, subtask = msg
                self.progress["value"] = progress
                self.status_label.config(text=status)
                self.subtask_label.config(text=subtask)
                self.root.update()
                
        except queue.Empty:
            if self.running:
                self.root.after(100, self.check_queue)

    def update_progress(self, progress, status, subtask=""):
        """Update the progress bar"""
        self.queue.put((progress, status, subtask))

    def start(self):
        """Start the progress bar"""
        self.root.mainloop()

    def stop(self):
        """Stop the progress bar"""
        self.running = False
        self.queue.put(None)
def run_script_with_progress():
    progress_bar = ProgressBar()
    
    def script_thread():
        try:

            # Initialize COM for this thread
            pythoncom.CoInitialize()
            # Step 1: Initialize browser setup
            progress_bar.update_progress(10, "Step 1/10: Setting up browser environment", "Configuring Playwright settings...(insert folder name)")
            time.sleep(2)  # Simulating work, replace with actuctual function calls
            def get_browser_path():
                try:
                    if getattr(sys, '_MEIPASS', False):
                        # Running as exe
                        return os.path.join(sys._MEIPASS, 'ms-playwright')
                    else:
                        # Running as script
                        return os.path.join(os.environ['LOCALAPPDATA'], 'ms-playwright')
                except Exception as e:
                    print(f"Error getting browser path: {e}")
                    return None

            # Set Playwright browser directory
            os.environ['PLAYWRIGHT_BROWSERS_PATH'] = get_browser_path()
            logger = logging.getLogger(__name__)
            # Step 2: Browser automation
            
            ##### Process Main Data tab #####
            downloads_path = os.path.join(os.path.expanduser("~"), "Downloads")
            folder_name= input("Insert Folder Name:")
            folder_path = os.path.join(downloads_path, folder_name)
            progress_bar.update_progress(20, f"Step 2/10 Navigating through {folder_name} folder")

            ### Grabing the week number from the folder name
            try:
                Week_num = int(folder_name.split(' ')[1])
                print(f'Got week number: {Week_num}')

                def get_week_start(week_number, year=2025):
                    # Create a date object for January 1st of the given year
                    jan_first = datetime.datetime(year, 1, 1)
                    
                    # Calculate the offset to the first week of the year
                    # If Jan 1 is before Thursday, it's considered part of week 1
                    # Otherwise, it's part of the last week of the previous year
                    if jan_first.weekday() <= 3:
                        first_week_start = jan_first - timedelta(days=jan_first.weekday())
                    else:
                        first_week_start = jan_first + timedelta(days=7 - jan_first.weekday())
                    
                    # Calculate the start of the desired week
                    target_date = first_week_start + timedelta(weeks=week_number-1)
                    
                    return target_date

                
                rep_day = get_week_start(Week_num,2025) - timedelta(days=1)
                rep_day = rep_day.date()
                print(f'Found Reporting Day: {rep_day}')
            except (IndexError, ValueError):
                print(f"Unable to extract week number from folder name: {folder_name}")
                Week_num = None
            ##### Process Main Data tab #####

            def print_ordered_files(folder_path, order_keywords=None):
                """
                Print files in order based on keywords in filename.
                
                Args:
                    folder_path: Path to the folder
                    order_keywords: List of keywords to order by
                """
                folder = Path(folder_path)
                
                # Default order if none provided
                if order_keywords is None:
                    order_keywords = ['raw', 'processed', 'final']

                # Get all files
                files = list(folder.glob('*'))
                
                # Create a dictionary to store files by category
                categorized_files = {keyword: [] for keyword in order_keywords}
                uncategorized = []

                # Categorize files based on keywords
                for file in files:
                    categorized = False
                    for keyword in order_keywords:
                        if keyword.lower() in file.name.lower():
                            categorized_files[keyword].append(file)
                            categorized = True
                            break
                    if not categorized:
                        uncategorized.append(file)

                # Print files in order
                print("\nFiles in ordered categories:")
                print("=" * 50)
            
                total_count = 0
                # Print categorized files
                dict ={}
                for keyword in order_keywords:
                    if categorized_files[keyword]:
                        print(f"\n{keyword.upper()} files:")
                        print("-" * 30)
                        for idx, file in enumerate(sorted(categorized_files[keyword]), 1):
                            total_count += 1
                            print(f"{total_count}. {file.name}")
                            name = file.name
                            
                        dict.update({keyword: file.name})

                return dict

            # Example usage with different ordering scenarios   
            data_order = ['Metrics Week Beginning', 'Turns Analysis Week Starting', 'Peel Pile','Hunt Dedicated', 'Driver Shift Detail']
            print("\nData processing ordering:")

            categorized_files=print_ordered_files(folder_path, data_order)


            print(categorized_files)
            ###### Define all paths ######
            try:
                Metrics = categorized_files['Metrics Week Beginning']
                metrics_path = os.path.join(folder_path, Metrics)
                if not os.path.exists(metrics_path):
                    progress_bar.update_progress(20, f"Metrics file not found: {metrics_path}")
                    raise FileNotFoundError(f"Metrics file not found: {metrics_path}")
            except FileNotFoundError as e:
                print(f"Warning: {e}")
                print("Skipping operations related to Metrics file.")
                progress_bar.update_progress(20, f"Metrics file not found: {metrics_path}.... Skipping operations")
                metrics_path = None

            try:
                main = categorized_files['Turns Analysis Week Starting']
                main_path = os.path.join(folder_path, main)
                if not os.path.exists(main_path):
                    progress_bar.update_progress(20,f"Main analysis file not found: {main_path}")
                    raise FileNotFoundError(f"Main analysis file not found: {main_path}")
            except FileNotFoundError as e:
                print(f"Warning: {e}")
                progress_bar.update_progress(20,f"Main analysis file not found: {main_path}.... Skipping operations")
                print("Skipping operations related to Main analysis file.")
                main_path = None


            # Now you can use metrics_path and main_path in your code
            if metrics_path:
                print(f"Processing Metrics file: {metrics_path}")
                progress_bar.update_progress(30,f"Step 3/10 processing metrics file: {metrics_path}")
                Location = pd.read_excel(metrics_path, sheet_name='Locations')
                Location = Location.iloc[:,0:6]
                Landing = pd.read_excel(metrics_path, sheet_name='Landing Page')
                Landing = Landing.iloc[0:9, 0:12]
                print(Landing)
                try:
                    Data = pd.read_excel(metrics_path, sheet_name='INV_AMZN_EDI_Detail')
                    Data = Data[(Data['Location ']=='LAX')&(Data['DCCount']==1)&(Data['Driver Shift']=='Electric')]
                    Trucks =Data[(Data['Location ']=='LAX')&(Data['Driver Shift']=='Electric')]
                    EDC_Trucks = [25000, 25001, 25002, 25003, 25004, 25005, 25006, 25007, 25008, 25009, 25010, 25011]
                    # Create a basic pivot table
                    pivot = pd.pivot_table(
                        data=Data,           # Your DataFrame
                        values='Driver',       # Values to aggregate
                        index=['EDC Truck','EDCShift'],           # Rows
                        columns='Date',      # Columns
                        aggfunc='count',          # Aggregation function
                        margins=True,            # Add row and column totals
                        margins_name='Total'     # Name for the total row/column
                    )

                    # Convert pivot table to long format DataFrame
                    # First, remove the margins (totals)
                    pivot_no_margins = pivot.drop('Total', axis=0).drop('Total', axis=1)

                    # Stack the columns to create a long format
                    df_long = pivot_no_margins.stack().reset_index()
                    df_long.columns = ['EDC Truck', 'EDCShift', 'Date', 'Driver Count']
                    df_long['EDC Truck'] = df_long['EDC Truck'].astype(int).astype(str)


                    # First, let's create a function to identify trucks with both AM and PM shifts on the same date
                    def has_am_pm_same_day(group):
                        # Check if the group has both 'AM' and 'PM' in the EDCShift column
                        shifts = group['EDCShift'].unique()
                        return 1 if ('AM' in shifts and 'PM' in shifts) else 0

                    # Group the data by EDC Truck and Date, then apply the function
                    result = df_long.groupby(['EDC Truck', 'Date']).apply(has_am_pm_same_day).reset_index()
                    result.columns = ['EDC Truck', 'Date', 'Has_AM_PM']

                    result = result.groupby('EDC Truck').agg({'Has_AM_PM': 'sum'}).reset_index()
                    result['EDC Truck'] = result['EDC Truck'].astype(int).astype(str)

                    df_long = df_long.groupby(['EDC Truck']).agg({'Date': 'nunique'}).reset_index()
                    df_long.rename(columns={'Date': 'Utilization'}, inplace=True)
                    df_long = df_long.merge(result, on='EDC Truck', how='left')
                    df_long.rename(columns={'Has_AM_PM': 'Slip Seats'}, inplace=True)

                    # Create a new DataFrame with just the EDC Truck column
                    truck_df = pd.DataFrame({'EDC Truck': EDC_Trucks})
                    truck_df['EDC Truck'] = truck_df['EDC Truck'].astype(int).astype(str)
                    print(truck_df)

                    # Merge with your existing DataFrame
                    df_long = truck_df.merge(df_long, on='EDC Truck', how='outer')
                    df_long['Slip Seats'] = df_long['Slip Seats'].fillna(0)
                    df_long['Utilization'] = df_long['Utilization'].fillna(0)
                    df_long['Slip Seats'] = df_long['Slip Seats'].astype(int)
                    df_long['Utilization'] = df_long['Utilization'].astype(int)

                    print(df_long)
                except:
                    print(f'Could not perform slip seat operation due to missing data')
                Data = pd.read_excel(metrics_path, sheet_name='INV_AMZN_EDI_Detail')
                def convert_mixed_time_column(df, column_name):
                    try:
                        # Time format pattern (HH:MM:SS)
                        time_pattern = re.compile(r'^\d{2}:\d{2}:\d{2}$')
                        
                        def process_value(value):
                            try:
                                # If already in correct format, return as is
                                if isinstance(value, str) and time_pattern.match(value):
                                    return value, None  # No change
                                
                                # Convert decimal to time
                                decimal = float(value)
                                decimal = decimal % 1 if decimal >= 1 else decimal
                                total_seconds = decimal * 24 * 60 * 60
                                hours = int(total_seconds // 3600)
                                minutes = int((total_seconds % 3600) // 60)
                                seconds = int(total_seconds % 60)
                                
                                new_value = f"{hours:02d}:{minutes:02d}:{seconds:02d}"
                                return new_value, f"Converted from {decimal:.9f}"
                                
                            except:
                                return value, None  # No change
                        
                        print("\nProcessing conversions...")
                        print("-" * 50)
                        
                        conversion_count = 0
                        # Process each value and track changes
                        for idx, value in df[column_name].items():
                            new_value, status = process_value(value)
                            df.at[idx, column_name] = new_value
                            if status:  # Only print if there was a change
                                #print(f"Row {idx}: {value} → {new_value} ({status})")
                                conversion_count += 1
                            
                        print("-" * 50)
                        print(f"Total conversions: {conversion_count}")
                        
                        return df
                        
                    except Exception as e:
                        print(f"Error during conversion: {str(e)}")
                        return df

                Data = convert_mixed_time_column(Data, 'Shift')
                def get_am_pm(time_str):
                    """
                    Convert time from 'HH:MM:SS' format to just AM or PM
                    Input: '10:19:00'
                    Output: 'AM' or 'PM'
                    """
                    try:
                        # Convert to string if it's not already
                        time_str = str(time_str)
                        hours = int(time_str.split(':')[0])
                        return 'PM' if 12 <= hours < 24 else 'AM'
                    except:
                        return time_str  # Return original value if conversion fails


                def convert_time_columns_to_ampm(df, time_columns):
                    """
                    Convert multiple time columns in a DataFrame to AM/PM
                    """
                    for col in time_columns:
                        new_col_name = "AM/PM"
                        df[new_col_name] = df[col].apply(get_am_pm)
                    return df

                Data = convert_time_columns_to_ampm(Data, ['Shift'])
                Data['Day of Week'] = pd.to_datetime(Data['Date'], errors='coerce').dt.day_name()
                Data = Data[['Location ', 'Date', 'Day of Week', 'Shift', 'Job Reference', 'Leg','Status', 'Container', 'CHS', 'Origin Zip','Origin', 'Destination Zip', 'Destination', 'Miles', 'Driver Shift', 'Origin Type','Origin Class', 'Dest Type', 'Dest Class', 'Driver','AM/PM','Stop Number', 'MBL']]
                Data =Data[(~Data['Location '].isin([0])) & (~Data['Location '].isna())]
                print(f'Stop Number column type before conversion: {Data["Stop Number"].dtype}')
                for value in Data['Stop Number']:
                    #print(f'Processing value: {value} of type {type(value)}')
                    if isinstance(value, str) and value.isdigit():
                        Data['Stop Number'] = Data['Stop Number'].replace(value, int(value))
                    elif isinstance(value, float) and value.is_integer():
                        Data['Stop Number'] = Data['Stop Number'].replace(value, int(value))
                    elif isinstance(value, int):
                        Data['Stop Number'] = Data['Stop Number'].replace(value, value)
                    elif isinstance(value, str):
                        try:
                            print(value)
                            Data['Stop Number'] = Data['Stop Number'].replace(value, -1)
                        except ValueError:
                            print(f"ValueError: Could not convert {value} to int, setting to -1")
                
                Data['Stop Number'] = Data['Stop Number'].fillna(-1)
                Data['Stop Number'] = Data['Stop Number'].astype(int)
                Data['Cont/MBL'] = Data.apply(lambda row: f"{row['Container']}{row['MBL']}" if pd.notna(row['Container']) and pd.notna(row['MBL']) else '', axis=1)
                try:
                    Data['If Slip?'] = np.where(Data['Driver Shift'].str.contains('Company DR Sft 3', na=False, case=False),'Slip Seat','')
                    print(f'Data conversion completed for slip seat')
                except Exception as e:
                    print(f"Error processing 'If Slip?' column: {e}")
                    Data['If Slip?'] = ''
                print(f'Data conversion completed for slip seat')
                Data['Is first/last?'] = np.where((Data['Stop Number'] == 0) | (Data['Stop Number']== 99),'Y','N')
                Data['Origin-Dest Class'] = Data.apply(lambda row: f"{row['Origin Class']}-{row['Dest Class']}" if pd.notna(row['Origin Class']) and pd.notna(row['Dest Class']) else '', axis=1)
                Data['Date']= pd.to_datetime(Data['Date']).dt.date
                Data['Year'] = pd.to_datetime(Data['Date']).dt.year
                Data['Week'] = pd.to_datetime(Data['Date']).dt.isocalendar().week
                Data['Week'] = np.where((Data['Week']== Week_num - 1)& (Data['Date'] == rep_day), Week_num, Data['Week'])
                Data['Week'] = np.where((Data['Day of Week']== 'Sunday')& (Data['Date'] != rep_day), Week_num-1, Data['Week'])
                Data['TL/FC'] = np.where((Data['Origin Class'].str.contains('TL', na=False, case=False)) | (Data['Dest Class'].str.contains('TL', na=False, case=False)), 'TL', 'FC')
                move_dict = {
                    'FC-PORT': 0.5,
                    'PORT-FC': 0.5,
                    'PORT-TL': 0.5,
                    'TL-PORT': 0.5,
                    'FC-YARD': 0.25,
                    'PORT-YARD': 0.25,
                    'TL-YARD': 0.25,
                    'YARD-FC': 0.25,
                    'YARD-PORT': 0.25,
                    'YARD-TL': 0.25,
                    'EXAM-FC': 0.375,
                    'EXAM-TL': 0.375,
                    'EXAM-YARD': 0.125,
                    'FC-EXAM': 0.125,
                    'PORT-EXAM': 0.125,
                    'EXAM-PORT': 0.125,
                    'TL-EXAM': 0.125,
                    'YARD-EXAM': 0.125,
                    '#N/A-#N/A': '',
                    '#N/A-TL': '',
                    '#N/A-YARD': '',
                    'CHSPROV-#N/A': '',
                    'CHSPROV-FC': '',
                    'CHSPROV-PORT': '',
                    'CHSPROV-TL': '',
                    'CHSPROV-YARD': '',
                    'EXAM-#N/A': '',
                    'FC-#N/A': '',
                    'PORT-#N/A': '',
                    'PORT-CHSPROV': '',
                    'TL-#N/A': '',
                    'TL-CHSPROV': '',
                    'YARD-#N/A': '',
                    'YARD-CHSPROV': '',
                    'EXAM-CHSPROV': '',
                    'CHSPROV-CHSPROV': '',
                    'FC-CHSPROV': '',
                    'CHSPROV-EXAM': '',
                    '#N/A-PORT': '',
                    'FC-FC': 0.125,
                    'FC-TL': 0.125,
                    'PORT-PORT': '',
                    'TL-FC': 0.125,
                    'TL-TL': 0.125,
                    'YARD-YARD': '',
                    '':''
                }
                Data['Weight with bob/chs for column AC']= Data['Origin-Dest Class'].map(move_dict)
                Data['Weight']= np.where((Data['Status'].str.contains('CHS', na=False, case=False)) | (Data['Status'].str.contains('BOB',na=False,case=False))," " ,Data['Weight with bob/chs for column AC'])
                Data['Driver/Week']= Data.apply(lambda row: f"{row['Week']}-{row['Driver']}" if pd.notna(row['Week']) and pd.notna(row['Driver']) else '', axis=1)
                # Convert non-string values to string first
                # Method 1: Using str.lower() with a condition
                Data['Job Reference'] = Data['Job Reference'].apply(lambda x: x.lower() if isinstance(x, str) else x)
                Data['Job Reference'] = Data['Job Reference'].astype(str)
                Data = Data[~Data['Job Reference'].str.contains('allocated driver', na=False)]
                def calculate_duplicate_reciprocal(df, column_name):
                    """
                    Replicates Excel formula =IFERROR(1/COUNTIFS($AG:$AG,AG2),"")
                    
                    Parameters:
                    df: pandas DataFrame
                    column_name: string, name of the column to check for duplicates
                    
                    Returns:
                    Series with reciprocal of duplicate counts or empty string if error
                    """
                    try:
                        # Count occurrences of each value
                        value_counts = df[column_name].value_counts()
                        
                        # Calculate reciprocal, handle division by zero with empty string
                        reciprocal = df[column_name].map(lambda x: round(1/value_counts[x],2) if pd.notna(x) else '')
                        
                        return reciprocal
                    
                    except Exception as e:
                        print(f"Error: {e}")
                        return pd.Series([''] * len(df))
                Data['Driver Weighted Moves'] = calculate_duplicate_reciprocal(Data, 'Driver/Week')
                Landing_dict= Landing[['Fleet','Total']]
                Landing_dict=Landing_dict.set_index('Fleet')['Total'].to_dict()
                Data['Weekly Driver Count']= Data['Location '].map(Landing_dict).astype(int)
                Control_T= Data[Data['Status'].isin(['FCL','ECL'])]
                float_containers = Control_T[Control_T['Container'].apply(lambda x: isinstance(x, float))]
                MBL = float_containers['MBL'].dropna().astype(str).tolist()
                file_name= 'Control_Tower_Quicksight.xlsx'
                path=os.path.join(os.path.expanduser('~'), 'Downloads', file_name)
                if MBL:  # This checks if the list is not empty
                    print(f'found MBL(s): {MBL}')
                else:
                    print("No MBL found")
                containers = Control_T['Container'].dropna().astype(str).tolist()
                print(f'found container(s): {len(containers)}')
                Containers = '\n'.join(containers)
                                # Basic version
                data1 = {
                    'Containers': Containers
                }

                # Save to file if needed
                with open(os.path.join(folder_path,'data1.json'), 'w') as f:
                    json.dump(data1, f, indent=2)

                ### GVT Container Search 
                def is_running_as_executable():
                    return getattr(sys, 'frozen', False) and hasattr(sys, '_MEIPASS')

                def get_base_path():
                    """Get the correct base path"""
                    if is_running_as_executable():
                        return sys._MEIPASS
                    return os.path.dirname(os.path.abspath(__file__))

                def find_node_executable():
                    """Find the Node.js executable"""
                    if is_running_as_executable():
                        base_path = get_base_path()
                        node_path = os.path.join(base_path,'node', 'node.exe')
                        
                        print(f"Running as executable. Looking for Node.js at: {node_path}")
                        
                        if os.path.exists(node_path):
                            return node_path
                        
                        print("\nContents of MEIPASS directory:")
                        for file in os.listdir(base_path):
                            print(f"  File: {file}")
                    else:
                        print("Running as script. Using system Node.js.")
                    
                    return 'node'  # Use system Node.js
                def setup_playwright_env():
                    """Setup Playwright environment"""
                    if getattr(sys, 'frozen', False):
                        if is_running_as_executable():
                            base_path = get_base_path()
                            print(base_path)
                        else:
                            base_path = os.path.join(os.path.expanduser("~"),'AppData','Local')
                            print(base_path)
                        playwright_path = os.path.join(base_path, 'ms-playwright')
                        print(playwright_path)
                        # Set environment variables
                        os.environ['PLAYWRIGHT_BROWSERS_PATH'] = playwright_path
                        os.environ['PLAYWRIGHT_SKIP_BROWSER_DOWNLOAD'] = '1'
                        
                        print(f"\nPlaywright environment:")
                        print(f"Browser path: {playwright_path}")
                        
                        # Verify browser files
                        if os.path.exists(playwright_path):
                            print("\nBrowser files:")
                            for root, dirs, files in os.walk(playwright_path):
                                level = root.replace(playwright_path, '').count(os.sep)
                                indent = ' ' * 4 * level
                                print(f"{indent}{os.path.basename(root)}/")
                                subindent = ' ' * 4 * (level + 1)
                                for f in files:
                                    print(f"{subindent}{f}")
                        
                        return True
                    return False

                def run_script(filename):
                    try:

                        if not setup_playwright_env():
                            print("Failed to setup Playwright environment")
                        base_path =get_base_path()

                        js_path = os.path.join(base_path, filename)
                        if not os.path.exists(js_path):
                            print(f"JavaScript file not found at: {js_path}")
                        node_path = find_node_executable()

                        print(f"\nUsing paths:")
                        print(f"Base path: {base_path}")
                        print(f"Node path: {node_path}")
                        print(f"JS path: {js_path}")

                        # Run the Node.js script
                        process = subprocess.Popen(
                            [node_path, js_path, folder_name],
                            cwd=base_path,
                            stdout=subprocess.PIPE,
                            stderr=subprocess.PIPE,
                            text=True,
                            bufsize=1,
                            universal_newlines=True,
                            env=os.environ
                        )

                        # Add immediate status check
                        print(f"Process ID: {process.pid}")
                        print(f"Process running: {process.poll() is None}")


                        def read_stream(stream):
                            output = []
                            while True:
                                line = stream.readline()
                                if not line:
                                    break
                                output.append(line)
                                print(line.strip())  # Immediate feedback
                            return output

                        with ThreadPoolExecutor(max_workers=2) as executor:
                            # Start reading both streams
                            stdout_future = executor.submit(read_stream, process.stdout)
                            stderr_future = executor.submit(read_stream, process.stderr)

                        # Get final status
                        print(f"\nExit code: {process.returncode}")
                        #print(f"Stdout: {stdout_future.result()}")
                        #print(f"Stderr: {stderr_future.result()}")

                        # Print output in real-time
                        while True:
                            output = process.stdout.readline()
                            if output == '' and process.poll() is not None:
                                break
                            if output:
                                print(output.strip())

                        # Check for errors
                        if process.returncode != 0:
                            print("Error running script:", process.stderr.read())
                        if  (len(stderr_future.result()) > 0) and (filename =='GVT_Search.js'):
                            pyperclip.copy(Containers)
                            print('Container list has been copied to your clipboard, please manually search GVT.')
                            file_name = input('input your GVT file name here:' )
                            return file_name

                    except Exception as e:
                        print(f"Error: {e}")
                        traceback.print_exc()
                        file_name = None 
                        return file_name

                    finally:
                        input("Press Enter to Continue...")

                if __name__ == "__main__":
                    output = input('Do you want to run a GVT search for these containers automatically? (Y/N):')
                    if output =='Y':
                        file_name = run_script('GVT_Search.js')
                    if output =='N':
                        pyperclip.copy(Containers)
                        print('Container list has been copied to your clipboard, please manually search GVT.')
                        file_name = input('input your GVT file name here:')
                
                progress_bar.update_progress(30,"Added Containers to your clipboard, please paste into GVT and input your GVT file name")
                ######### GVT download ###########
                if file_name is None:
                    file_name = f'GVT Search File {datetime.date.today().strftime('%m.%d.%y')}'
                path=os.path.join(folder_path, f'{file_name}'+'.xlsx')
                print(f'this is your file path: {path}')

                GVT_Data= pd.read_excel(path)
                GVT_Data = GVT_Data.iloc[:,0:77]
                if len([x for x in GVT_Data['Container'].tolist() if x not in containers]) == 0:
                    print('GVT file contains more containers than the original search. Please check your GVT file.')
                    file_name = input('input your GVT file name here:' )
                    path=os.path.join(folder_path, f'{file_name}'+'.xlsx')
                    print(f'this is your file path: {path}')
                    GVT_Data= pd.read_excel(path)
                    GVT_Data = GVT_Data.iloc[:,0:77]
                GVT_Data['Container2'] = GVT_Data['Container'].tolist()
                
                def create_type_column(df):
                    # List of facilities that should be marked as "FBA FCL- AFTX"
                    fba_facilities = ['XHH3', 'XSE2', 'XLA4', 'MSKW-LKW2', 'XPH1', 'XEW3']
                    
                    # Create new Type column using numpy.where (similar to IF statement)
                    df['Type'] = np.where(df['Facility'].isin(fba_facilities), 
                                        'FBA FCL- AFTX', 
                                        df['Category'])
                    
                    return df
                GVT_Data = create_type_column(GVT_Data)
                GVT_Data= GVT_Data.sort_values('Estimated Pickup',ascending=False)
                GVT_Data= GVT_Data.drop_duplicates(subset=['Container'], keep='first').reset_index(drop=True)
                def handle_managing_facility(df):
                    """
                    Ensures the Managing Facility column exists and is positioned as the 4th column in the dataframe.
                    
                    Parameters:
                    df (DataFrame): The GVT dataframe to modify
                    
                    Returns:
                    DataFrame: Modified dataframe with Managing Facility as the 4th column
                    """
                    # First check if Managing Facility exists and handle any None values
                    try:
                        # If column exists, fill NA values with empty string
                        df['Managing Facility'] = df['Managing Facility'].fillna('')
                    except KeyError:
                        # If column doesn't exist, create it with empty strings
                        df['Managing Facility'] = ''
                    
                    # Get all column names
                    cols = df.columns.tolist()
                    
                    # If Managing Facility is already in the dataframe, remove it from its current position
                    if 'Managing Facility' in cols:
                        cols.remove('Managing Facility')
                    
                    # Insert Managing Facility as the 4th column (index 3)
                    # Make sure we have at least 3 columns before inserting
                    if len(cols) >= 3:
                        cols.insert(3, 'Managing Facility')
                    else:
                        # If we have fewer than 3 columns, add it at the end
                        cols.append('Managing Facility')
                    
                    # Reorder the dataframe columns
                    df = df[cols]
                    
                    return df
                # Ensure Managing Facility exists and is the 4th column
                GVT_Data = handle_managing_facility(GVT_Data)
                GVT_Data_cols= GVT_Data[['Container','Container Size','Container Priority Code','Type']]
                GVT_Data_cols['Relay if value']= ' '
                GVT_Data_cols['Query lookup']= ' '
                GVT_Data_cols['Priority Calc'] = np.where(GVT_Data_cols['Container Priority Code'].str.contains('EXPRESS', na=False, case=False), 'Express', '')
                GVT_Data_cols['CD or TL'] = np.where(GVT_Data_cols['Type'].str.contains('Retail Transload', na=False, case=False) |GVT_Data_cols['Type'].str.contains('FBA FCL- AFTX',na=False,case=False),'TL','CD')
                cols = GVT_Data_cols.columns.tolist()
                cols.remove('Relay if value')
                cols.remove('Query lookup')
                cols.remove('Priority Calc')
                cols.insert(2, 'Relay if value')
                cols.insert(3, 'Query lookup')
                cols.insert(5,'Priority Calc')
                GVT_Data_cols = GVT_Data_cols[cols]
                Data = pd.merge(Data, GVT_Data_cols, right_on= 'Container', left_on='Container', how='left')
                print(Data)
            else:
                progress_bar.update_progress(30,f"Metrics file not found {Metrics}.. skipping step")
                pass
            ###### Peel Pile ######
            try:
                Pile = categorized_files['Peel Pile']
                peel_path = os.path.join(folder_path, Pile)
                progress_bar.update_progress(40,f"Step 4/10 processing peel pile file {peel_path}")
                if not os.path.exists(peel_path):
                    raise FileNotFoundError(f"Peel Pile file not found: {peel_path}")
            except FileNotFoundError as e:
                progress_bar.update_progress(40,f"Error {e} Skipping operations related to Peel Pile file.")
                print(f"Warning: {e}")
                print("Skipping operations related to Peel Pile file.")
                peel_path = None
            if peel_path:
                print(f"Peel Pile file found: {peel_path}")
                old_peels = pd.read_excel(main_path, sheet_name='Peel Pile')
                peel_pile = pd.read_excel(peel_path)
                peels= peel_pile['Pickup Number'].tolist()
                old_peels = old_peels[~old_peels['Pickup Number'].isin(peels)]
                peel_pile = pd.concat([old_peels,peel_pile], axis = 0)
                peel_pile['Week']= pd.to_datetime(peel_pile['PU_STAGE']).dt.isocalendar().week
                peel_cols = peel_pile[['Container', 'Dock Status']]
                peel_cols = peel_cols.drop_duplicates(subset=['Container'],keep='first')
                print(peel_cols)
                Data = pd.merge(Data, peel_cols, right_on='Container', left_on='Container', how='left')
                Data['Peel Calc']= np.where(Data['Dock Status'].str.contains('PEEL PILE', na=False, case=False)& Data['Status'].str.contains('FCL',na=False, case=False), 'Peel Pile', ' ')
                print(Data)
            else:
                progress_bar.update_progress(40,f"Error {e} Skipping operations related to Peel Pile file.")
                pass
            ##### Missed containers - Control Tower #####
            if metrics_path:
                GVT_Containers = GVT_Data['Container'].dropna().astype(str).tolist()
                progress_bar.update_progress(50,f'Step 5/10 processing Control Tower Containers')
                Control_cont = [container for container in containers if container not in GVT_Containers]
                print(f'GVT missed these Containers: {Control_cont}')
                MBL_grab = Data[Data['Container'].isin(Control_cont)]
                MBL_grab = MBL_grab['MBL'].dropna().astype(str).tolist()
                MBL = MBL_grab + MBL
                MBL_list = [x for x in MBL if pd.notna(x)]
                print(f'found MBLs: {MBL_list}')

                # Basic version
                data = {
                    'MBL': '\n'.join(MBL_list),
                    'Containers': '\n'.join(Control_cont)
                }
                downloads_path = os.path.join(os.path.expanduser("~"),"Downloads", folder_name)
                # Save to file if needed
                with open(os.path.join(downloads_path,'data.json'), 'w') as f:
                    json.dump(data, f, indent=2)
                if __name__ == "__main__":
                    run_script('Hunt_Data_Collection.js')
                if os.path.exists(f"{downloads_path}/Control_Tower_Data {datetime.date.today().strftime('%m.%d.%y')}.xlsx"):
                    print(f"Found File Control Tower Data {rep_day.strftime('%m.%d.%y')}.xlsx")    
                    control= pd.read_excel(f'{downloads_path}'+f'/Control_Tower_Data {datetime.date.today().strftime('%m.%d.%y')}.xlsx', header= 2)
                    control = control[['container_id','bol','booking_id','business','dray_type','pod']]
                    control['Cont/MBL'] = control.apply(lambda row: f"{row['container_id']}{row['bol']}" if pd.notna(row['container_id']) and pd.notna(row['bol']) else '', axis=1)
                    missing_cont= control['container_id'].tolist()
                    print(f'missing container(s):{missing_cont}')
                    Control_MBL = control['bol'].tolist()
                    print(f'The MBL list is {MBL_list}')
                    Missing_MBLs = [i for i in MBL_list if i not in Control_MBL]
                    print(f'Control tower did not find: {Missing_MBLs}')
                    my_dict = dict(zip(Control_MBL, missing_cont))
                    for index, row in Data[Data['Status'].isin(['FCL','ECL'])].iterrows():
                        if row['MBL'] in my_dict:
                            Data.at[index, 'Container'] = my_dict[row['MBL']]
                            #print(f'Row {index}: MBL: {row['MBL']}, Container {row['Container']} to {my_dict[row['MBL']]}')
                    lookup_dict = dict(zip(control['container_id'],control['dray_type']))
                    Data['Control Tower Lookup']= Data['Container'].map(lookup_dict)
                    cols = Data.columns.tolist()
                    cols.remove('Control Tower Lookup')
                    cols.insert(38,'Control Tower Lookup')
                    Data = Data[cols]
                    Data['CD or TL'] = np.where(Data['Control Tower Lookup'].str.contains('AFTX', na=False, case=False), 'TL', Data['CD or TL'])
                    Data['CD or TL']= np.where(Data['Control Tower Lookup'].str.contains('CONTAINER_DIRECT', na=False, case=False), 'CD', Data['CD or TL'])
                    Data['CD or TL'] = np.where(Data['Location '].str.contains('SAV', na=False, case=False) & Data['Type'].str.contains('AMAZON ROBOTICS', na=False, case=False), 'TL', Data['CD or TL'])
                    print(Data)
                    print(peel_pile)
                else:
                    print(f"Did not find File Control_Tower_Data {datetime.date.today().strftime('%m.%d.%y')}.xlsx.... Skipping operations")
                    Data['Control Tower Lookup']= ''
                    cols = Data.columns.tolist()
                    cols.remove('Control Tower Lookup')
                    cols.insert(38,'Control Tower Lookup')
                    Data = Data[cols]
                    Data['CD or TL'] = np.where(Data['Control Tower Lookup'].str.contains('AFTX', na=False, case=False), 'TL', Data['CD or TL'])
                    Data['CD or TL']= np.where(Data['Control Tower Lookup'].str.contains('CONTAINER_DIRECT', na=False, case=False), 'CD', Data['CD or TL'])
                    Data['CD or TL'] = np.where(Data['Location '].str.contains('SAV', na=False, case=False) & Data['Type'].str.contains('AMAZON ROBOTICS', na=False, case=False), 'TL', Data['CD or TL'])
                    print(Data)
                    print(peel_pile)

            else:
                progress_bar.update_progress(50,'Skipping Step 5 no metrics data')
                pass
            #### Volvo Data Sharepoint ####
            try:
                progress_bar.update_progress(90, f"Step 9/10: Processing final data updates")
                Volvo = pd.read_excel(f'{downloads_path}'+f'/2025_Project_Shazam_WBR {datetime.date.today().strftime('%m.%d.%y')}.xlsx',sheet_name=f'Wk{Week_num}')
                Volvo = Volvo.iloc[1:14,0:12]
                # Get the name of the 7th column
                column_to_drop = Volvo.columns[6]
                Volvo[' ']= ' '
                # Drop the 7th column
                Volvo = Volvo.drop(columns=[column_to_drop])
                cols= Volvo.columns.tolist()
                cols.remove(' ')
                cols.insert(6, ' ')
                Volvo = Volvo[cols]
            except ValueError as e:
                print(f'Error: {e}')

            #### Allocations ####
            try:
                Allocations = categorized_files['Hunt Dedicated']
                allocations_path = f'{folder_path}/{Allocations}'
                progress_bar.update_progress(60,f'Step 6/10 Processing allocations data {allocations_path}')
                if not os.path.exists(allocations_path):
                    raise FileNotFoundError(f"Allocation file not found: {allocations_path}")
            except FileNotFoundError as e:
                print(f"Warning: {e}")
                progress_bar.update_progress(60,f'Error file not found{allocations_path}... skipping step 6')
                print("Skipping operations related to Allocation file.")
                allocations_path = None
            if allocations_path:
                print(f"Allocations file found: {allocations_path}")
                Allocations = pd.read_excel(allocations_path, sheet_name='Hunt Volume',engine ='pyxlsb')
                Allocations = Allocations.iloc[11:22,4:30]
                print(Allocations)
                # Make first row as headers and remove that row
                new_header = Allocations.iloc[0]  # Grab first row for header
                Allocations = Allocations[1:]  # Take data less the header row
                Allocations.columns = new_header  # Set new header row
                Allocations.reset_index(drop=True, inplace=True)
                Allocations = Allocations.rename(columns={Allocations.columns[0]: ' '})
                def sort_dataframe_by_custom_list(df, column, custom_order):
                    """
                    Sort a DataFrame by a custom list order for a specific column.
                    
                    Parameters:
                    df (DataFrame): The pandas DataFrame to sort
                    column (str): The column to sort by
                    custom_order (list): List of values in the desired sort order
                    
                    Returns:
                    DataFrame: Sorted DataFrame
                    """
                    # Create a categorical data type with our custom order
                    cat_type = pd.CategoricalDtype(categories=custom_order, ordered=True)
                    
                    # Convert the column to the categorical type
                    df[column] = df[column].astype(cat_type)
                    
                    # Sort the DataFrame
                    sorted_df = df.sort_values(column)
                    
                    # Convert back to original type if needed
                    # df[column] = df[column].astype(str)  # Uncomment if needed
                    
                    return sorted_df

                # Define the custom order for the locations
                custom_location_order = [' ','LAX', 'ORF','OAK','CHI', 'SAV', 'EWR', 'SEA', 'TIW', 'PNW', 'Total']
                # Fix the error in your code - use the function instead of sorted()
                Allocations = Allocations[Allocations[' '].isin(custom_location_order)]
                Allocations = sort_dataframe_by_custom_list(Allocations, ' ', custom_location_order)
                Allocations = Allocations[[' ', 'Planned Drivers', 'Utilized', '% Drivers Utilized', 'Backlog', 
                                        'Current Week Arrivals', 'TL', 'FC', '% of 1P port volume', 
                                        'Rollover into next week (Thur-Sun)', 'Notes']]
                print(Allocations)
            else:
                progress_bar.update_progress(60,f'Error file not found {allocations_path}... skipping step 6')
                pass

            ### Driver Shifts #####
            try:
                shift_data = categorized_files['Driver Shift Detail']
                shift_path = f'{folder_path}/{shift_data}'
                if not os.path.exists(shift_path):
                    raise FileNotFoundError(f"Driver Shift file not found: {shift_path}")
            except FileNotFoundError as e:
                print(f"Warning: {e}")
                print("Skipping operations related to Driver Shift file.")
                shift_path = None
            if shift_path:
                print(f"Driver Shift file found: {shift_path}")
                progress_bar.update_progress(70,f'Step 7/10 processing driver shift data {shift_path}')
                shift_hours = pd.read_excel(shift_path)
                shift_hours= shift_hours.iloc[:-2,:]
                # Convert all datetime columns to strings to avoid timezone issues
                for col in shift_hours.columns:
                    if pd.api.types.is_datetime64_dtype(shift_hours[col]):
                        shift_hours[col] = shift_hours[col].astype(str)
                port_mapping = {
                    "AMAZON PORT - CHIIL": "CHI",
                    "AMAZON PORT - NEWNJ": "EWR",
                    "AMAZON PORT - CMPCA": "LAX",
                    "AMAZON PORT - FONCA": "LAX",
                    "AMAZON PORT - OAKCA": "OAK",
                    "AMAZON PORT - NORVA": "ORF",
                    "AMAZON PORT - SAVGA": "SAV",
                    "AMAZON PORT - SEAWA": "PNW",
                    "AMAZON PORT - TACWA": "PNW",
                    "AMAZON PORT - PNWWA": "PNW"
                }
                shift_hours['port_map']= shift_hours['AccountName'].map(port_mapping)
                # Get all column names
                cols = shift_hours.columns.tolist()
                cols.remove('port_map')
                cols.insert(0,'port_map')
                shift_hours = shift_hours[cols]
                # Add a flag column for records over 14 hours (1) or under 4 hours (2)
                def get_hours_flag(hours):
                    if hours > 14:
                        return 2  # Flag for over 14 hours
                    elif hours < 4:
                        return 1  # Flag for under 4 hours
                    else:
                        return 0  # Normal hours
                # Apply the function to create the flag column
                shift_hours['Hours Flag'] = shift_hours['Hours (Adj)'].apply(get_hours_flag)
                #print("Added 'Hours Flag' column: 1 = over 14 hours, 2 = under 4 hours, 0 = normal hours")
                # Define the locations you want to include
                locations = ['CHI', 'EWR', 'LAX', 'OAK', 'ORF', 'SAV', 'PNW']

                # Filter out records with hours > 14
                filtered_hours = shift_hours[shift_hours['Hours (Adj)'] <= 14]

                # Calculate average hours by location
                avg_hours_by_location = filtered_hours.groupby('port_map')['Hours (Adj)'].mean()

                # Create dictionary with all desired locations, using 0 for missing locations
                avg_hours_dict = {location: float(round(avg_hours_by_location.get(location, 0), 2)) for location in locations}

                for index, row in shift_hours[shift_hours['Hours (Adj)'] >=14].iterrows():
                    if row['port_map'] in avg_hours_dict:
                        shift_hours.at[index, 'Hours (Adj)'] = avg_hours_dict[row['port_map']]
                        print(f'Row {index}:  {row['Hours (Adj)']} hours to {avg_hours_dict[row['port_map']]} hours')
                print(shift_hours)
            else:
                progress_bar.update_progress(70,f'Skipping step 7 no driver shift data {shift_path}')
                pass


            def convert_xlsm_to_xlsx(input_path, output_path=None):
                excel = None
                wb = None
                try:
                    # Initialize COM
                    pythoncom.CoInitialize()
                    
                    # Input validation
                    if not os.path.exists(input_path):
                        raise FileNotFoundError(f"Input file not found: {input_path}")
                    
                    # Ensure input file is .xlsm
                    if not input_path.lower().endswith('.xlsm'):
                        raise ValueError(f"Input file must be .xlsm format: {input_path}")
                    
                    # Create output path if not provided
                    if output_path is None:
                        input_file = Path(input_path)
                        output_path = input_file.with_stem(input_file.stem + "_no_macro").with_suffix('.xlsx')
                    
                    # Start Excel with proper error handling
                    excel = win32com.client.DispatchEx("Excel.Application")
                    excel.Visible = False
                    excel.DisplayAlerts = False
                    
                    try:
                        # Open the .xlsm workbook
                        wb = excel.Workbooks.Open(str(os.path.abspath(input_path)))
                        
                        if wb is None:
                            raise Exception("Failed to open workbook")
                        
                        # Save as .xlsx
                        full_output_path = str(os.path.abspath(output_path))
                        wb.SaveAs(Filename=full_output_path, FileFormat=51)  # 51 = xlOpenXMLWorkbook (no macros)
                        
                    except Exception as wb_error:
                        logger.error(f"Workbook operation failed: {str(wb_error)}")
                        raise
                    
                    finally:
                        # Close workbook if it was opened
                        if wb is not None:
                            try:
                                wb.Close(SaveChanges=False)
                            except:
                                pass

                    logger.info(f"Successfully converted '{input_path}' to standard workbook: '{output_path}'")
                    return output_path

                except Exception as e:
                    logger.error(f"Error converting file: {str(e)}")
                    return None

                finally:
                    # Cleanup Excel
                    if excel is not None:
                        try:
                            excel.Quit()
                        except:
                            pass
                    
                    # Release COM objects
                    try:
                        if wb is not None:
                            del wb
                        if excel is not None:
                            del excel
                    except:
                        pass
                    
                    # Uninitialize COM
                    try:
                        pythoncom.CoUninitialize()
                    except:
                        pass


            def batch_convert_folder(folder_path, output_folder=None, rep_day=None, Week_Num=None):
                """
                Convert all .xlsm files in a folder to .xlsx and add week/date information to filename
                
                Parameters:
                folder_path (str): Path to folder containing .xlsm files
                output_folder (str): Path to output folder (optional)
                rep_day (str): Reporting day in format MM.DD.YY
                Week_Num (str/int): Week number
                """
                if not os.path.exists(folder_path):
                    raise FileNotFoundError(f"Folder not found: {folder_path}")
                
                # Create output folder if specified
                if output_folder and not os.path.exists(output_folder):
                    os.makedirs(output_folder)
                
                # Get all .xlsm files in the folder
                xlsm_files = [f for f in os.listdir(folder_path) if f.endswith('.xlsm')]
                
                for file in xlsm_files:
                    input_path = os.path.join(folder_path, file)
                    
                    # Create new filename with week and date information
                    file_name = Path(file).stem  # Get filename without extension

                    rep_day = rep_day.strftime('%m.%d.%y')
                    
                    # Extract year from rep_day if provided (assuming format MM.DD.YY)
                    year = f"20{rep_day[-2:]}" if rep_day else "2025"  # Default to 2025 if no rep_day
                    
                    # Construct new filename
                    if Week_Num and rep_day:
                        new_file_name = f"{year} Week {Week_Num} Turns Analysis Week Starting {rep_day}.xlsx"
                    else:
                        new_file_name = f"{file_name} (copy).xlsx"
                    
                    if output_folder:
                        output_path = os.path.join(output_folder, new_file_name)
                    else:
                        output_path = os.path.join(folder_path, new_file_name)
                        
                    print(f"\nProcessing: {file}")
                    print(f"New filename will be: {new_file_name}")
                    
                    converted_path = convert_xlsm_to_xlsx(input_path, output_path)
                    if converted_path:
                        print(f"Converted: {converted_path}")
                    else:
                        print(f"Failed to convert: {file}")
            # Example usage:
            if __name__ == "__main__":
                try:
                    batch_convert_folder(folder_path,rep_day=rep_day, Week_Num=Week_num)
                    progress_bar.update_progress(80,f'Step 8/10 Converting {main_path} to xlsx')
                    print("Batch conversion completed!")
                except Exception as e:
                    print(f"Batch conversion failed: {str(e)}")

            if main_path:
                print(f"Main file found: {main_path}")
                # Read the WBR Rollup sheet
                WBR_long =pd.read_excel(main_path, sheet_name='WBR File')
                WBR_long = WBR_long.iloc[:153,2:6]
                print(WBR_long)

                WBR_data = pd.read_excel(main_path, sheet_name='WBR Rollup')
                WBR_weeks = WBR_data.iloc[2:24,7:13]
                # Make first row as headers and remove that row
                new_header = WBR_weeks.iloc[0]  # Grab first row for header
                WBR_weeks = WBR_weeks[1:]  # Take data less the header row
                WBR_weeks.columns = new_header  # Set new header row
                WBR_weeks.reset_index(drop=True, inplace=True)
                print(WBR_weeks)
                WBR_YTD = WBR_data.iloc[15:24, [14]]
                print(WBR_YTD)
                EV_Data_YTD = WBR_data.iloc[28:42,[13]]
                print(EV_Data_YTD)
            else:
                
                pass
            def extract_total_invoice_cost(dataframe):
                total_cost = None
                for index, row in dataframe.iterrows():
                    if any("Total Invoice Cost:" in str(cell) for cell in row):
                        total_cost = row.iloc[1]  # Get value from the 10th column (index 1)
                        break
                return total_cost

            def process_excel_file(file_path, sheet_names):
                result_dict = {}
                
                try:
                    # Get all sheet names in the Excel file
                    excel_sheets = pd.ExcelFile(file_path)
                    available_sheets = excel_sheets.sheet_names
                    
                    # Iterate through specified sheets
                    for sheet_name in sheet_names:
                        if sheet_name in available_sheets:
                            # Read the sheet into a DataFrame
                            df = pd.read_excel(
                                file_path, 
                                sheet_name=sheet_name,
                                engine='openpyxl'  # or 'xlrd' for older .xls files
                            )
                            df = df.iloc[:, [1, -1]]

                            # Extract the total invoice cost
                            total_cost = extract_total_invoice_cost(df)
                            
                            # Add to result dictionary if a cost was found
                            if total_cost is not None:
                                sheet_key = sheet_name[-3:]
                                result_dict[sheet_key] = total_cost
                        else:
                            print(f"Warning: Sheet '{sheet_name}' not found in the workbook.")
                    
                except Exception as e:
                    print(f"Error processing Excel file: {str(e)}")
                
                return result_dict
            # Usage

            sheets_to_process = ['Key Performace Indicators - CHI', 'Key Performace Indicators - EWR', 'Key Performace Indicators - LAX','Key Performace Indicators - ORF','Key Performace Indicators - SAV','Key Performace Indicators - PNW']

            if metrics_path:
                print(f"Metrics file found: {metrics_path}")
                results = process_excel_file(metrics_path, sheets_to_process)
                GM_NA = pd.read_excel(f'{downloads_path}'+f'/GM_NA_Carrier_Data {datetime.date.today().strftime('%m.%d.%y')}.xlsx', header =5)
                GM_NA = GM_NA.loc[[0, 1, 2, 3, 4, 5], ['Region', Week_num]]

                Naming_dict = { 'USCHI':'CHI','USLAX':'LAX','USNYC':'EWR','USORF':'ORF','USPNW':'PNW','USSAV':'SAV'}
                GM_NA['Region'] = GM_NA['Region'].map(lambda x: Naming_dict.get(x, x))
                GM_NA['Invoices'] = GM_NA['Region'].map(results)
                print(GM_NA)
            else:
                GM_NA = None
                pass
            def clear_range_before_paste(ws, start_row, start_col, df, include_header, include_index):
                """
                Clear a range in the worksheet based on DataFrame dimensions
                """
                try:
                    # Calculate dimensions
                    num_rows = len(df)
                    num_cols = len(df.columns)
                    total_rows = num_rows + (1 if include_header else 0)
                    total_cols = num_cols + (1 if include_index else 0)
                    end_row = start_row + total_rows - 1
                    end_col = start_col + total_cols - 1

                    # Get the range to clear
                    clear_range = ws.Range(
                        ws.Cells(start_row, start_col),
                        ws.Cells(end_row, end_col)
                    )

                    print(f"Clearing range: {get_column_letter(start_col)}{start_row}:"
                        f"{get_column_letter(end_col)}{end_row}")
                    
                    # Clear contents and formatting
                    clear_range.ClearContents()
                    clear_range.ClearFormats()
                    
                    return True
                except Exception as e:
                    print(f"Error clearing range: {str(e)}")
                    return False
            # Set up logging
            logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
            logger = logging.getLogger(__name__)

            def clear_sheet(ws):
                """
                Clears sheet contents based on specific rules for different sheet names
                
                Args:
                    ws: Excel worksheet object
                """
                pythoncom.CoInitialize()
                try:
                    logger.info(f"Clearing sheet: {ws.Name}")
                    # Get the used range
                    last_row = ws.UsedRange.Rows.Count
                    last_col = ws.UsedRange.Columns.Count
                    logger.info(f"Sheet dimensions - Last Row: {last_row}, Last Column: {last_col}")

                    start_row = 1
                    start_col = 1

                    # Use if/elif structure to ensure only one condition is executed
                    if ws.Name == 'WBR File':
                        logger.info(f"Preserving first row and clearing from row 4")
                        clear_range = ws.Range(
                            ws.Cells(start_row + 3, start_col + 4),
                            ws.Cells(last_row, last_col-13)
                        )
                        range_address = f'{get_column_letter(start_col+4)}{start_row+3}:{get_column_letter(last_col-13)}{last_row}'
                        print(f"Clearing range for WBR File sheet: {range_address}")
                        clear_range.ClearContents()
                    elif ws.Name == 'GVT Data':
                        logger.info(f"Preserving first row and last four columns for GVT data")
                        clear_range = ws.Range(
                            ws.Cells(start_row+1, start_col),
                            ws.Cells(last_row, last_col - 2)
                        )
                        range_address = f'{get_column_letter(start_col)}{start_row+1}:{get_column_letter(last_col-4)}{last_row}'
                        clear_range.ClearContents()

                    elif ws.Name == 'Data Tab w calcs':
                        logger.info(f"Preserving first row")
                        clear_range = ws.Range(
                            ws.Cells(start_row + 1, start_col),
                            ws.Cells(last_row, last_col)
                        )
                        range_address = f'{get_column_letter(start_col)}{start_row+1}:{get_column_letter(last_col)}{last_row}'
                        clear_range.ClearContents()

                    elif ws.Name == 'Allocations':
                        logger.info(f"Preserving first five rows")
                        clear_range = ws.Range(
                            ws.Cells(start_row + 4, start_col + 1),
                            ws.Cells(last_row, last_col)
                        )
                        range_address = f'{get_column_letter(start_col+1)}{start_row+4}:{get_column_letter(last_col)}{last_row}'
                        clear_range.ClearContents()

                    elif ws.Name == 'Shift Hours':
                        logger.info(f'Preserving first thirteen rows')
                        clear_range = ws.Range(
                            ws.Cells(start_row + 13, start_col),
                            ws.Cells(last_row, last_col)
                        )
                        range_address = f'{get_column_letter(start_col)}{start_row+13}:{get_column_letter(last_col)}{last_row}'
                        clear_range.ClearContents()

                    elif ws.Name == 'Drivers':
                        logger.info(f'Preserving first fourteen rows and clearing multiple ranges')
                        try:
                            # Clear first range
                            clear_range1 = ws.Range(
                                ws.Cells(start_row + 14, start_col),
                                ws.Cells(20, 12)
                            )
                            clear_range1.ClearContents()
                            
                            # Clear second range
                            clear_range2 = ws.Range(
                                ws.Cells(start_row + 24, start_col + 1),
                                ws.Cells(37, 3)
                            )
                            clear_range2.ClearContents()
                            
                            # Clear third range
                            clear_range3 = ws.Range(
                                ws.Cells(start_row + 25, 15),
                                ws.Cells(38, 25)
                            )
                            clear_range3.ClearContents()
                            
                            range_address = "Multiple ranges cleared"
                        except Exception as e:
                            logger.error(f"Error clearing Drivers sheet ranges: {str(e)}")
                            raise

                    elif ws.Name == 'Control Tower Lookup':
                        logger.info('Preserving first row')
                        clear_range = ws.Range(
                            ws.Cells(start_row + 1, start_col),
                            ws.Cells(last_row, last_col)
                        )
                        range_address = f'{get_column_letter(start_col)}{start_row+1}:{get_column_letter(last_col)}{last_row}'
                        clear_range.ClearContents()

                    else:
                        logger.info(f"Clearing entire used range")
                        clear_range = ws.UsedRange
                        range_address = clear_range.Address
                        clear_range.ClearContents()

                        logger.info(f'Successfully cleared sheet range: {range_address}')
                    return True

                except Exception as e:
                    logger.error(f"Error in clear_sheet for {ws.Name}: {str(e)}")
                    logger.debug(traceback.format_exc())
                    return False

                finally:
                    pythoncom.CoUninitialize()
                    try:
                        ws.Application.ScreenUpdating = True
                    except:
                        pass

            def safe_paste_data_to_excel(data_dict, file_path, formatting=None, 
                                    date_format=None,
                                    include_index=False, clear_before_paste=True,
                                    fast_sheets=None, clear_from_row=None,
                                    preserve_formula_sheets=None,
                                    dont_shift_data=False ):
                """
                Safely pastes data to Excel using Win32COM with advanced formatting.
                """
                pythoncom.CoInitialize()
                if not data_dict:
                    print("No data to process. Exiting function.")
                    return False
                excel = None
                wb = None

                excel = win32com.client.Dispatch("Excel.Application")
                time.sleep(1)
                excel.Visible = True
                excel.DisplayAlerts = False
                
                file_path = Path(file_path)
                backup_file_path = file_path.with_name(f"{file_path.stem}_backup{file_path.suffix}")

                def convert_to_excel_date(value):
                    """
                    Convert Python date/time objects to Excel-compatible values
                    """
                    try:
                        if isinstance(value, datetime.datetime):
                            return value.strftime('%Y-%m-%d %H:%M:%S')
                        elif isinstance(value, datetime.date):
                            return value.strftime('%Y-%m-%d')
                        elif isinstance(value, pd.Timestamp):
                            return value.strftime('%Y-%m-%d %H:%M:%S')
                        elif isinstance(value, datetime.time):
                            return value.strftime('%H:%M:%S')
                        else: 
                            return value
                    except Exception as e:
                        print(f"Error converting value {value}: {str(e)}")
                        return value

                def prepare_data_for_excel(df, include_header, include_index):
                    """
                    Prepare DataFrame data for Excel, handling date conversions and formatting
                    """
                    try:
                        # Convert DataFrame to a list of lists
                        if include_index:
                            data = df.reset_index().values.tolist()
                        else:
                            data = df.values.tolist()
                        
                        # Add header if required
                        if include_header:
                            header = df.reset_index().columns.tolist() if include_index else df.columns.tolist()
                            data.insert(0, header)
                        
                        # Convert dates and other types
                        for row in data:
                            for i, value in enumerate(row):
                                if pd.isna(value):
                                    row[i] = None
                                else:
                                    row[i] = convert_to_excel_date(value)
                        
                        return data
                    except Exception as e:
                        print(f"Error preparing data: {str(e)}")
                        return []
                def clear_sheet_contents(ws, start_row, preserve_formatting=True):
                    """
                    Clear sheet contents from specified row
                    """
                    try:
                        last_row = ws.UsedRange.Rows.Count
                        last_col = ws.UsedRange.Columns.Count
                        
                        print(f"Clearing data from row {start_row} to {last_row}")
                        clear_range = ws.Range(
                            ws.Cells(start_row, 1),
                            ws.Cells(last_row, last_col)
                        )
                        
                        if preserve_formatting:
                            clear_range.ClearContents()  # Preserves formatting
                        else:
                            clear_range.Clear()  # Clears everything including formatting
                            
                        return True
                    except Exception as e:
                        print(f"Error clearing sheet contents: {str(e)}")
                        return False

                def apply_shift_hours_formatting(ws, df, start_row, start_col):
                    """
                    Apply conditional formatting to Shift Hours sheet
                    """
                    try:
                        # Find the flag column index (assuming 'Hours Flag' is the column name)
                        flag_col_name = 'Hours Flag'
                        if flag_col_name not in df.columns:
                            print(f"Warning: '{flag_col_name}' column not found in DataFrame")
                            return False

                        flag_col_idx = df.columns.get_loc(flag_col_name)
                        flag_col = start_col + flag_col_idx + (1 if include_index else 0)
                        end_row = start_row + len(df)
                        end_col = start_col + len(df.columns) - 1

                        print(f"Applying conditional formatting based on Hours Flag column ({flag_col})")
                        
                        # Process each row
                        for row_idx, flag_value in enumerate(df[flag_col_name], start=start_row + (1 if include_header else 0)):
                            # Define the range for this row
                            row_range = ws.Range(
                                ws.Cells(row_idx, start_col),
                                ws.Cells(row_idx, end_col)
                            )
                            
                            # Clear existing formatting
                            row_range.Interior.Color = -4142  # Excel's default background
                            
                            # Apply new formatting based on flag value
                            if flag_value == 2:  # Under 4 hours
                                row_range.Interior.Color = 65535  # Yellow
                                print(f"Row {row_idx}: Applied yellow formatting (flag value 2)")
                            elif flag_value == 1:  # Over 14 hours
                                row_range.Interior.Color = 255  # Red
                                print(f"Row {row_idx}: Applied red formatting (flag value 1)")

                        return True
                    except Exception as e:
                        print(f"Error applying shift hours formatting: {str(e)}")
                        return False
                try:
                    # Create backup of original file if it exists
                    if file_path.exists():
                        print(f"Creating backup: {backup_file_path}")
                        shutil.copy2(file_path, backup_file_path)
                        wb = excel.Workbooks.Open(str(file_path))
                    else:
                        wb = excel.Workbooks.Add()
                        wb.SaveAs(str(file_path))
                    

                    time.sleep(1)  # Allow Excel to stabilize

                
                    fast_sheets = set(fast_sheets or [])

                    preserve_formula_sheets = set(preserve_formula_sheets or [])
                    # At the beginning of your function, after opening the workbook:
                    worksheet_dict = {sheet.Name: sheet for sheet in wb.Worksheets}

                    # Then replace the nested loops with:
                    for sheet_name in preserve_formula_sheets:
                        if sheet_name in worksheet_dict:
                            ws = worksheet_dict[sheet_name]
                            clear_sheet(ws)

                            

                    if dont_shift_data:
                        data_dict.pop('WBR Rollup')
                        data_dict.pop('WBR File')
                    for sheet_name, df_list in data_dict.items():
                        # Skip if df_list is empty
                        if not df_list:
                            print(f"Skipping empty df_list for sheet: {sheet_name}")
                            continue
                        print(f"\nProcessing sheet: {sheet_name}")
                        
                        # Check if sheet exists
                        sheet_exists = False
                        for sheet in wb.Worksheets:
                            if sheet.Name == sheet_name:
                                sheet_exists = True
                                ws = sheet
                                break
                        
                        if not sheet_exists:
                            print(f"Creating new sheet: {sheet_name}")
                            ws = wb.Worksheets.Add()
                            ws.Name = sheet_name
                        
                        ws.Activate()

                        # Clear contents for fast sheets if specified
                        if sheet_name in fast_sheets and clear_from_row:
                            print(f"Clearing existing content in fast sheet {sheet_name}")
                            clear_sheet_contents(ws, clear_from_row, preserve_formatting=True)

                        for df, start_row, start_col, include_header in df_list:
                            # Skip if DataFrame is empty
                            if df is None or df.empty:
                                print(f"Skipping empty DataFrame at position ({start_row}, {start_col})")
                                continue
                            print(f'Pasting df:\n{df.head()}')

                            if GM_NA is not None and sheet_name == "Finance Cost WBR Input":
                                ws.Activate()
                                used_range = ws.UsedRange

                                regions_found = set()
                                regions_not_found = set(df['Region'])

                                # Iterate through rows in the first column
                                for row in range(1, used_range.Rows.Count + 1):
                                    cell_value = ws.Cells(row, 1).Value
                                    
                                    if cell_value in df['Region'].values:
                                        print(f"Found region: {cell_value}")
                                        print(f'{cell_value} is {type(cell_value)} dtype')    
                                        regions_found.add(cell_value)
                                        regions_not_found.remove(cell_value)
                                        
                                        row_data = df[df['Region'] == cell_value].iloc[0]
                                        print(row_data)
                                        # Set target row based on region using string comparison
                                        if str(cell_value).strip() == 'LAX':
                                            target_row = row + 23
                                        elif str(cell_value).strip() == 'PNW':
                                            target_row = row + 20
                                        elif str(cell_value).strip() == 'EWR':
                                            target_row = row + 16
                                        elif str(cell_value).strip() == 'SAV':
                                            target_row = row + 16
                                        elif str(cell_value).strip() == 'ORF':
                                            target_row = row + 16
                                        elif str(cell_value).strip() == 'CHI':
                                            target_row = row + 13
                                        else:
                                            logger.warning(f"Unrenrecognized region: {cell_value}")
                                            continue

                                        for column_name, value in row_data.items():
                                            if column_name != 'Region':
                                                try:
                                                    ws.Cells(target_row + 1, 2).Value = value
                                                    target_row += 1
                                                except Exception as e:
                                                    logger.error(f"Error pasting data for {cell_value}, column {column_name}: {str(e)}")

                                        logger.info(f"Data pasted horizontally for region: {cell_value}")

                                if regions_not_found:
                                    logger.warning(f"Regions not found in Excel: {', '.join(regions_not_found)}")
                            else:
                            
                                # Clear specific range if needed (for non-fast sheets or when clear_from_row isn't specified)
                                if clear_before_paste and (sheet_name not in fast_sheets or not clear_from_row):
                                    clear_range = ws.Range(
                                        ws.Cells(start_row, start_col),
                                        ws.Cells(start_row + len(df) + (1 if include_header else 0), 
                                                start_col + len(df.columns) + (1 if include_index else 0))
                                    )
                                    clear_range.ClearContents()

                                # Prepare data
                                data_to_paste = prepare_data_for_excel(df, include_header, include_index)

                                # Fast pasting method for designated sheets
                                if sheet_name in fast_sheets:
                                    paste_range = ws.Range(
                                        ws.Cells(start_row, start_col),
                                        ws.Cells(start_row + len(data_to_paste) - 1, 
                                                start_col + len(data_to_paste[0]) - 1)
                                    )
                                    paste_range.Value = data_to_paste

                                    # Apply date formatting to date columns
                                    if date_format:
                                        for col_idx, col_name in enumerate(df.columns):
                                            if pd.api.types.is_datetime64_any_dtype(df[col_name]):
                                                col_range = ws.Range(
                                                    ws.Cells(start_row + (1 if include_header else 0), start_col + col_idx),
                                                    ws.Cells(start_row + len(data_to_paste) - 1, start_col + col_idx)
                                                )
                                                col_range.NumberFormat = date_format

                                    # Apply basic formatting
                                    if formatting and include_header:
                                        header_range = ws.Range(
                                            ws.Cells(start_row, start_col),
                                            ws.Cells(start_row, start_col + len(data_to_paste[0]) - 1)
                                        )
                                        if 'header' in formatting:
                                            header_format = formatting['header']
                                            if 'bold' in header_format and header_format['bold']:
                                                header_range.Font.Bold = True
                                            if 'color' in header_format:
                                                header_range.Font.Color = int(header_format['color'], 16)
                                            if 'bg_color' in header_format:
                                                header_range.Interior.Color = int(header_format['bg_color'], 16)
                                        # Special handling for Finance Cost WBR Input sheet

                                else:
                                    # Standard pasting method
                                    for row_idx, row in enumerate(data_to_paste):
                                        for col_idx, value in enumerate(row):
                                            cell = ws.Cells(start_row + row_idx, start_col + col_idx)
                                            cell.Value = value

                                            # Apply formatting
                                            if row_idx == 0 and include_header and formatting and 'header' in formatting:
                                                # Apply header formatting
                                                header_format = formatting['header']
                                                if 'bold' in header_format and header_format['bold']:
                                                    cell.Font.Bold = True
                                            
                                            # Apply date formatting if needed
                                            if isinstance(value, str) and any(c in value for c in [':', '-']):
                                                try:
                                                    datetime.datetime.strptime(value, '%Y-%m-%d')
                                                    cell.NumberFormat = date_format or "yyyy-mm-dd"
                                                except ValueError:
                                                    pass

                        # Apply conditional formatting for Shift Hours sheet
                        if sheet_name == 'Shift Hours':
                            print("Applying Shift Hours specific formatting...")
                            apply_shift_hours_formatting(ws, df, start_row, start_col)

                    print("\nSaving workbook...")
                    wb.Save()
                    print("Operation completed successfully!")
                    return True

                except Exception as e:
                    print(f"Error in safe_paste_data_to_excel: {str(e)}")
                    if backup_file_path.exists():
                        print("Restoring from backup...")
                        shutil.copy2(backup_file_path, file_path)
                    return False

                finally:
                    pythoncom.CoUninitialize()
                    try:
                        if wb:
                            wb.Close(SaveChanges=True)
                        if excel:
                            excel.ScreenUpdating = True
                            excel.Quit()
                        if backup_file_path.exists():
                            os.remove(backup_file_path)
                    except:
                        pass

            # Prepare your data dictionary with multiple DataFrames per sheet
            data_dict = {}

            # WBR File
            if 'WBR_long' in locals() or 'WBR_long' in globals():
                data_dict['WBR File'] = [(WBR_long, 1, 2, True)]

            # WBR Rollup
            WBR_list = []
            if 'WBR_weeks' in locals() or 'WBR_weeks' in globals():
                WBR_list.append((WBR_weeks, 4, 6, True))
            if 'WBR_YTD' in locals() or 'WBR_YTD' in globals():
                WBR_list.append((WBR_YTD, 17, 21, False))
            if 'EV_Data_YTD' in locals() or 'EV_Data_YTD' in globals():
                WBR_list.append((EV_Data_YTD, 30, 21, False))
            if WBR_list:  # Only add if list is not empty
                data_dict['WBR Rollup'] = WBR_list

            # Data Tab w calcs
            if 'Data' in locals() or 'Data' in globals():
                data_dict['Data Tab w calcs'] = [(Data, 2, 1, False)]

            # Allocations
            if 'Allocations' in locals() or 'Allocations' in globals():
                data_dict['Allocations'] = [(Allocations, 5, 1, False)]

            # Shift Hours
            if 'shift_hours' in locals() or 'shift_hours' in globals():
                data_dict['Shift Hours'] = [(shift_hours, 14, 1, False)]

            # Drivers
            drivers_list = []
            if 'Landing' in locals() or 'Landing' in globals():
                drivers_list.append((Landing, 15, 1, False))
            if 'df_long' in locals() or 'df_long' in globals():
                drivers_list.append((df_long, 25, 1, False))
            if 'Volvo' in locals() or 'Volvo' in globals():
                drivers_list.append((Volvo,26,14,False))
            if drivers_list:  # Only add if list is not empty
                data_dict['Drivers'] = drivers_list

            # Peel Pile
            if 'peel_pile' in locals() or 'peel_pile' in globals():
                data_dict['Peel Pile'] = [(peel_pile, 2, 1, False)]

            # GVT Data
            if 'GVT_Data' in locals() or 'GVT_Data' in globals():
                data_dict['GVT Data'] = [(GVT_Data, 2, 1, False)]

            # Control Tower Lookup
            if 'control' in locals() or 'control' in globals():
                data_dict['Control Tower Lookup'] = [(control, 2, 1, False)]

            # Finance Cost WBR Input
            if 'GM_NA' in locals() or 'GM_NA' in globals():
                data_dict['Finance Cost WBR Input'] = [(GM_NA, 1, 1, False)]
            today = datetime.date.today()
            year, week_num, _ = today.isocalendar()
            date_obj = rep_day.strftime('%m.%d.%y')

            new_path = os.path.join(folder_path, f"{year} Week {Week_num} Turns Analysis Week Starting {date_obj}" + '.xlsx')
            # Call the function
            output = input('Shift WBR data (Y/N):')
            output = f'{output}' 
            if output =='Y':
                output=False
            else:
                output=True
            progress_bar.update_progress(95, f"Step 10/10: Saving final data to Excel")
            safe_paste_data_to_excel(
                data_dict,
                new_path, 
                clear_before_paste=False,
                fast_sheets=['WBR File','Data Tab w calcs', 'Shift Hours', 'Peel Pile', 'GVT Data','Control Tower Lookup'],
                preserve_formula_sheets=['WBR File','Data Tab w calcs', 'GVT Data','Control Tower Lookup','Shift Hours','Drivers','Allocations'],
                dont_shift_data= output
            )
            def refresh_excel_workbook(file_path):
                """
                Opens an Excel workbook from the given path and refreshes all data connections.
                
                Args:
                    file_path (str or Path): Path to the Excel workbook
                
                Returns:
                    bool: True if refresh was successful, False otherwise
                """
                pythoncom.CoInitialize()
                excel = None
                wb = None
                
                try:
                    print(f"Opening workbook: {file_path}")
                    # Convert to Path object if string
                    file_path = Path(file_path)
                    
                    # Check if file exists
                    if not file_path.exists():
                        print(f"Error: File not found at {file_path}")
                        return False
                        
                    # Initialize Excel with different method
                    try:
                        excel = win32com.client.DispatchEx("Excel.Application")
                    except Exception as e:
                        print(f"Error creating Excel application: {str(e)}")
                        excel = win32com.client.Dispatch("Excel.Application")
                        
                    # Set Excel properties safely
                    try:
                        excel.Visible = True
                    except:
                        print("Could not set Excel visibility - continuing anyway")
                        
                    try:
                        excel.DisplayAlerts = False
                    except:
                        print("Could not disable alerts - continuing anyway")
                    
                    # Open workbook
                    print("Opening Excel workbook...")
                    wb = excel.Workbooks.Open(str(file_path))
                    
                    # Attempt to refresh
                    print("Attempting to refresh workbook...")
                    try:
                        # Try direct refresh first
                        print("Attempting direct refresh...")
                        wb.RefreshAll()
                        time.sleep(30)  # Wait for refresh to complete
                        print("Direct refresh completed")
                        success = True
                    except Exception as e:
                        print(f"Direct refresh failed: {str(e)}")
                        try:
                            # Fallback to alternate refresh method
                            print("Trying alternate refresh method...")
                            excel.Application.Calculate()
                            excel.Application.CalculateUntilAsyncQueriesComplete()
                            time.sleep(30)
                            print("Alternate refresh completed")
                            success = True
                        except Exception as e2:
                            print(f"Alternate refresh failed: {str(e2)}")
                            success = False
                    
                    # Save after refresh
                    if success:
                        try:
                            print("Saving workbook...")
                            wb.Save()
                            print("Workbook saved successfully")
                        except Exception as e:
                            print(f"Error saving workbook: {str(e)}")
                            success = False
                    
                    return success
                    
                except Exception as e:
                    print(f"Error in refresh_excel_workbook: {str(e)}")
                    return False
                    
                finally:
                    if wb:
                        try:
                            wb.Close(SaveChanges=True)
                        except:
                            pass
                    if excel:
                        try:
                            excel.Quit()
                        except:
                            pass
                        finally:
                            pythoncom.CoUninitialize()
                            excel = None
            # Usage example:
            if __name__ == "__main__":
                # Example usage
                success = refresh_excel_workbook(new_path)
                if success:
                    print("Workbook refreshed successfully")
                else:
                    print("Failed to refresh workbook")

            # Grab new recent week and Data and paste it
            # read the Excel file
            df = pd.read_excel(new_path, sheet_name ='WBR Rollup')
            current_week = df.iloc[1:155,[1]]
            new_header = current_week.iloc[1]  # Grab first row for header
            current_week = current_week[2:]  # Take data less the header row
            current_week.columns = new_header  # Set new header row
            current_week.reset_index(drop=True, inplace=True)
            print(current_week)

            output= input('Paste new week data(Y/N):')
            if output =='Y':
                # Convert current_week to a DataFrame
                progress_bar.update_progress(98, "Pasting new week data...")
                safe_paste_data_to_excel(
                    {'WBR File': [(current_week, 3, 6, True)]},
                    new_path,
                    clear_before_paste=False)
                progress_bar.update_progress(100, "Process completed successfully!")
            else:
                print(f'Skipped Week {Week_num} paste to WBR File:')
                print('Process Completed')
                progress_bar.update_progress(100, f"Skipped Week {Week_num} paste to WBR File. Process completed!")
            def convert_to_macro_workbook(input_path, output_path=None):
                excel = None
                wb = None
                try:
                    # Initialize COM
                    pythoncom.CoInitialize()
                    
                    # Input validation
                    if not os.path.exists(input_path):
                        raise FileNotFoundError(f"Input file not found: {input_path}")
                        
                    # Create output path if not provided
                    if output_path is None:
                        input_file = Path(input_path)
                        output_path = input_file.with_stem(input_file.stem).with_suffix('.xlsm')
                    
                    # Start Excel with proper error handling
                    excel = win32com.client.DispatchEx("Excel.Application")
                    excel.Visible = True
                    excel.DisplayAlerts = True
                    
                    # Open and handle workbook properly
                    try:
                        wb = excel.Workbooks.Open(str(os.path.abspath(input_path)))
                        
                        # Ensure the workbook is actually open
                        if wb is None:
                            raise Exception("Failed to open workbook")
                            
                        # Save as .xlsm with full path
                        full_output_path = str(os.path.abspath(output_path))
                        wb.SaveAs(Filename=full_output_path, FileFormat=52)  # 52 = xlOpenXMLWorkbookMacroEnabled
                        
                    except Exception as wb_error:
                        logger.error(f"Workbook operation failed: {str(wb_error)}")
                        raise
                        
                    finally:
                        # Close workbook if it was opened
                        if wb is not None:
                            try:
                                wb.Close(SaveChanges=False)
                            except:
                                pass

                    logger.info(f"Successfully converted '{input_path}' to macro-enabled workbook: '{output_path}'")
                    return output_path

                except Exception as e:
                    logger.error(f"Error converting file: {str(e)}")
                    return None

                finally:
                    # Cleanup Excel
                    if excel is not None:
                        try:
                            excel.Quit()
                        except:
                            pass
                        
                    # Release COM objects
                    try:
                        if wb is not None:
                            del wb
                        if excel is not None:
                            del excel
                    except:
                        pass
                        
                    # Uninitialize COM
                    try:
                        pythoncom.CoUninitialize()
                    except:
                        pass

            # Usage with better error handling
            def safe_convert_workbook(input_path, output_path=None):
                try:
                    # Ensure no existing Excel processes are hanging
                    kill_excel_processes()
                    
                    # Add small delay to ensure system resources are released
                    time.sleep(1)
                    
                    # Attempt conversion
                    result = convert_to_macro_workbook(input_path, output_path)
                    
                    if result:
                        # Verify the file exists and is accessible
                        if os.path.exists(result):
                            return result
                        else:
                            raise FileNotFoundError("Output file was not created")
                    else:
                        raise Exception("Conversion returned None")
                        
                except Exception as e:
                    logger.error(f"Error during script execution: {str(e)}")
                    raise
                    
            # Helper function to kill hanging Excel processes
            def kill_excel_processes():
                try:
                    for proc in psutil.process_iter(['name']):
                        if proc.info['name'] == 'EXCEL.EXE':
                            try:
                                proc.kill()
                            except:
                                pass
                    time.sleep(1)  # Give system time to clean up
                except:
                    pass

            # Actual usage
            try:
                output_path = safe_convert_workbook(new_path)
                print(f"Successfully converted workbook to: {output_path}")
            except Exception as e:
                print(f"Failed to convert workbook: {str(e)}")
                # Handle failure appropriately
            progress_bar.update_progress(100,"Script completed successfully")
            logger.info("Script completed successfully")
        except Exception as e:
            error_msg = f"Error during script execution: {str(e)}"
            logger.error(error_msg)
            progress_bar.update_progress(0, "Error", error_msg)
        finally:
            progress_bar.stop()

    # Start the script in a separate thread
    thread = threading.Thread(target=script_thread)
    thread.start()

    # Start the progress bar
    progress_bar.start()

    # Wait for the script to complete
    thread.join()

if __name__ == "__main__":
    run_script_with_progress()


